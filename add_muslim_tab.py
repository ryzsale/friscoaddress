#!/usr/bin/env python3
"""Combine all Frisco zip code CSVs, filter Muslim names, and export to Excel."""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from muslim_filter import is_muslim_name

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
HEADERS     = ["Last Name", "First Name", "Address", "City / State / ZIP"]
FIELDS      = ["last_name", "first_name", "address", "city_state_zip"]
WIDTHS      = [22, 24, 38, 28]

FRISCO_ZIPS = ["75033", "75034", "75035", "75036"]
OUTPUT      = "frisco_all.xlsx"


def style_header(ws):
    for col_idx, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def append_row(ws, row_num, rec):
    ws.append([rec[f] for f in FIELDS])
    if row_num % 2 == 0:
        for cell in ws[row_num]:
            cell.fill = ALT_FILL


def address_sort_key(rec):
    addr = rec["address"].strip()
    parts = addr.split(None, 1)
    if len(parts) == 2 and parts[0].isdigit():
        return (parts[1].upper(), int(parts[0]))
    return (addr.upper(), 0)


def load_csv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# Load and deduplicate across zip codes (key: address + last_name + first_name)
seen = set()
all_records = []
muslim_records = []

for z in FRISCO_ZIPS:
    path = f"frisco_{z}.csv"
    try:
        rows = load_csv(path)
    except FileNotFoundError:
        # Fall back to test_output.csv for 75035 if frisco_75035.csv missing
        if z == "75035":
            rows = load_csv("test_output.csv")
        else:
            print(f"  Warning: {path} not found, skipping.")
            continue

    added = 0
    for rec in rows:
        key = (rec["address"].strip().upper(), rec["last_name"].lower(), rec["first_name"].lower())
        if key in seen:
            continue
        seen.add(key)
        all_records.append(rec)
        if is_muslim_name(rec["last_name"], rec["first_name"]):
            muslim_records.append(rec)
        added += 1
    print(f"  {path}: {added} unique records loaded")

muslim_records.sort(key=address_sort_key)

print(f"\nTotal unique owners: {len(all_records)}")
print(f"Muslim households:   {len(muslim_records)}")

# Build workbook
wb = Workbook()

ws_all = wb.active
ws_all.title = "All Owners"
style_header(ws_all)
for i, rec in enumerate(all_records, start=2):
    append_row(ws_all, i, rec)

ws_muslim = wb.create_sheet("Muslim Houses")
style_header(ws_muslim)
for i, rec in enumerate(muslim_records, start=2):
    append_row(ws_muslim, i, rec)

wb.save(OUTPUT)
print(f"\nSaved → {OUTPUT}")
