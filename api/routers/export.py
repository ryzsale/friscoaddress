import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..auth import get_current_user
from ..database import get_db
from ..data import get_dataframe
from ..models import Annotation
from ..routers.owners import _apply_filters

router = APIRouter(prefix="/export", tags=["export"])

HEADERS = ["Last Name","First Name","Address","Street","City/State/ZIP","Status","Last Visited","Comments"]
WIDTHS  = [20, 22, 36, 28, 26, 16, 16, 40]


def _build_rows(df, db: Session):
    ann_map = {a.rid: a for a in db.query(Annotation).filter(Annotation.rid.in_(df["rid"].tolist())).all()}
    rows = []
    for rec in df.to_dict("records"):
        a = ann_map.get(rec["rid"])
        rows.append({
            "Last Name":      rec["last_name"],
            "First Name":     rec["first_name"],
            "Address":        rec["address"],
            "Street":         rec["street"],
            "City/State/ZIP": rec["city_state_zip"],
            "Status":         a.status       if a else "",
            "Last Visited":   a.last_visited if a else "",
            "Comments":       a.comments     if a else "",
        })
    return rows


@router.get("/csv")
def export_csv(
    q: str = Query(""), zip: str = Query(""), muslim: str = Query(""),
    status: str = Query(""), street: str = Query(""),
    db: Session = Depends(get_db), _: str = Depends(get_current_user),
):
    df   = _apply_filters(get_dataframe().copy(), q, zip, muslim, status, False, street, db)
    rows = _build_rows(df, db)
    buf  = io.StringIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=frisco_export.csv"})


@router.get("/xlsx")
def export_xlsx(
    q: str = Query(""), zip: str = Query(""), muslim: str = Query(""),
    status: str = Query(""), street: str = Query(""),
    db: Session = Depends(get_db), _: str = Depends(get_current_user),
):
    df   = _apply_filters(get_dataframe().copy(), q, zip, muslim, status, False, street, db)
    rows = _build_rows(df, db)

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")

    wb = Workbook()
    ws = wb.active
    ws.title = "Frisco Export"
    ws.append(HEADERS)
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    for rn, r in enumerate(rows, 2):
        ws.append([r[h] for h in HEADERS])
        if rn % 2 == 0:
            for c in ws[rn]:
                c.fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=frisco_export.xlsx"})
