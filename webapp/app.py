#!/usr/bin/env python3
"""Frisco Address Admin — Flask web backend."""

import hashlib
import io
import os
import re
import sqlite3
import sys
from urllib.parse import quote_plus

import pandas as pd
from flask import (Flask, render_template, request, redirect,
                   url_for, send_file, flash, jsonify)
from flask_login import (LoginManager, UserMixin, login_user,
                         logout_user, login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Path setup ────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "annotations.db")
sys.path.insert(0, BASE_DIR)
from muslim_filter import is_muslim_name

app = Flask(__name__)
app.secret_key = "frisco-admin-secret-2024"
app.jinja_env.filters["urlencode"] = quote_plus

# ── Auth ──────────────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view = "login"

USERS = {"admin": generate_password_hash("admin123")}

class User(UserMixin):
    def __init__(self, username):
        self.id = username

@login_manager.user_loader
def load_user(user_id):
    return User(user_id) if user_id in USERS else None

# ── SQLite annotations ────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS annotations (
                rid          TEXT PRIMARY KEY,
                last_visited TEXT DEFAULT '',
                comments     TEXT DEFAULT '',
                status       TEXT DEFAULT '',
                ignored      INTEGER DEFAULT 0
            )
        """)
        conn.commit()

def record_id(rec):
    key = f"{str(rec.get('last_name','')).lower()}|{str(rec.get('first_name','')).lower()}|{str(rec.get('address','')).strip().upper()}"
    return hashlib.md5(key.encode()).hexdigest()[:16]

def get_annotations(rids):
    if not rids:
        return {}
    with get_db() as conn:
        ph   = ",".join("?" * len(rids))
        rows = conn.execute(f"SELECT * FROM annotations WHERE rid IN ({ph})", rids).fetchall()
    return {r["rid"]: dict(r) for r in rows}

def get_all_annotations():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM annotations").fetchall()
    return {r["rid"]: dict(r) for r in rows}

def upsert_annotation(rid, field, value):
    if field not in {"last_visited", "comments", "status", "ignored"}:
        return
    with get_db() as conn:
        conn.execute(f"""
            INSERT INTO annotations (rid, {field}) VALUES (?, ?)
            ON CONFLICT(rid) DO UPDATE SET {field}=excluded.{field}
        """, (rid, value))
        conn.commit()

def bulk_upsert(rids, field, value):
    if field not in {"status", "ignored"}:
        return
    with get_db() as conn:
        for rid in rids:
            conn.execute(f"""
                INSERT INTO annotations (rid, {field}) VALUES (?, ?)
                ON CONFLICT(rid) DO UPDATE SET {field}=excluded.{field}
            """, (rid, value))
        conn.commit()

# ── Data loading ──────────────────────────────────────────────────────────────
def extract_street(address):
    """Return the street name portion, e.g. '5408 BELLE CHASSE LN' → 'BELLE CHASSE LN'."""
    addr = address.strip().upper()
    m = re.match(r'^\d+\s+(.*)', addr)
    return m.group(1) if m else addr

def load_all_data():
    zip_files = [
        os.path.join(BASE_DIR, "frisco_75033.csv"),
        os.path.join(BASE_DIR, "frisco_75034.csv"),
        os.path.join(BASE_DIR, "frisco_75036.csv"),
        os.path.join(BASE_DIR, "test_output.csv"),
    ]
    frames = [pd.read_csv(p, dtype=str).fillna("") for p in zip_files if os.path.exists(p)]
    if not frames:
        return pd.DataFrame(columns=["last_name","first_name","address","city_state_zip"])
    df = pd.concat(frames, ignore_index=True)
    df = df.drop_duplicates(subset=["address","last_name","first_name"])
    df["is_muslim"] = df.apply(lambda r: is_muslim_name(r["last_name"], r["first_name"]), axis=1)
    df["zip"]       = df["city_state_zip"].str.extract(r'(\d{5})')
    df["street"]    = df["address"].apply(extract_street)
    df["rid"]       = df.apply(record_id, axis=1)
    return df

init_db()
DF = load_all_data()

# ── Constants ─────────────────────────────────────────────────────────────────
STATUS_OPTIONS = ["", "Muslim", "Non-Muslim", "Not to Visit"]

# ── Helpers ───────────────────────────────────────────────────────────────────
def apply_filters(df, q, zip_f, muslim_f, status_f, show_ignored, street_f=""):
    if q:
        mask = (
            df["last_name"].str.contains(q, case=False, na=False) |
            df["first_name"].str.contains(q, case=False, na=False) |
            df["address"].str.contains(q, case=False, na=False)
        )
        df = df[mask]
    if zip_f:
        df = df[df["zip"] == zip_f]
    if street_f:
        df = df[df["street"].str.contains(street_f, case=False, na=False)]
    if muslim_f == "yes":
        df = df[df["is_muslim"]]
    elif muslim_f == "no":
        df = df[~df["is_muslim"]]
    if status_f or not show_ignored:
        all_ann = get_all_annotations()
        if status_f:
            df = df[df["rid"].apply(lambda r: all_ann.get(r, {}).get("status", "") == status_f)]
        if not show_ignored:
            df = df[df["rid"].apply(lambda r: not all_ann.get(r, {}).get("ignored", 0))]
    return df

def enrich(slice_df):
    ann = get_annotations(slice_df["rid"].tolist())
    records = []
    for rec in slice_df.to_dict("records"):
        a = ann.get(rec["rid"], {})
        rec["last_visited"] = a.get("last_visited", "")
        rec["comments"]     = a.get("comments", "")
        rec["status"]       = a.get("status", "")
        rec["ignored"]      = bool(a.get("ignored", 0))
        records.append(rec)
    return records

def make_gmaps_route(addresses):
    """Build a Google Maps multi-stop directions URL (max 10 stops)."""
    stops = addresses[:10]
    encoded = [quote_plus(a) for a in stops]
    return "https://www.google.com/maps/dir/" + "/".join(encoded) + "/"

# ── Routes ────────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def index():
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@login_required
def dashboard():
    total      = len(DF)
    muslim     = int(DF["is_muslim"].sum())
    non_muslim = total - muslim

    all_ann = get_all_annotations()
    status_counts = {"Muslim": 0, "Non-Muslim": 0, "Not to Visit": 0}
    visited = ignored = 0
    for a in all_ann.values():
        s = a.get("status", "")
        if s in status_counts:
            status_counts[s] += 1
        if a.get("last_visited"):
            visited += 1
        if a.get("ignored"):
            ignored += 1

    by_zip = (DF.groupby("zip")["is_muslim"]
                .agg(total="count", muslim="sum")
                .reset_index()
                .sort_values("total", ascending=False))

    return render_template("dashboard.html",
        total=total, muslim=muslim, non_muslim=non_muslim,
        status_counts=status_counts, visited=visited,
        ignored=ignored, zip_stats=by_zip.to_dict("records"),
    )


@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        u = request.form.get("username","").strip()
        p = request.form.get("password","")
        if u in USERS and check_password_hash(USERS[u], p):
            login_user(User(u))
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/streets")
@login_required
def streets():
    zip_f    = request.args.get("zip", "").strip()
    muslim_f = request.args.get("muslim", "yes")   # default: Muslim only
    q        = request.args.get("q", "").strip()

    df = DF.copy()
    if zip_f:
        df = df[df["zip"] == zip_f]
    if muslim_f == "yes":
        df = df[df["is_muslim"]]
    elif muslim_f == "no":
        df = df[~df["is_muslim"]]

    # Exclude ignored
    all_ann = get_all_annotations()
    df = df[df["rid"].apply(lambda r: not all_ann.get(r, {}).get("ignored", 0))]

    # Visited count per street
    visited_rids = {rid for rid, a in all_ann.items() if a.get("last_visited")}

    grouped = (df.groupby(["street","zip"])
                 .agg(total=("rid","count"),
                      muslim=("is_muslim","sum"),
                      rids=("rid", list))
                 .reset_index())

    # Add visited count per street
    grouped["visited"] = grouped["rids"].apply(
        lambda rs: sum(1 for r in rs if r in visited_rids)
    )
    grouped = grouped.sort_values("muslim", ascending=False)

    if q:
        grouped = grouped[grouped["street"].str.contains(q, case=False, na=False)]

    street_list = grouped.to_dict("records")

    return render_template("streets.html",
        streets=street_list,
        zip_f=zip_f, muslim_f=muslim_f, q=q,
        zip_options=sorted(DF["zip"].dropna().unique()),
        total_streets=len(street_list),
        total_homes=int(grouped["muslim"].sum()) if len(grouped) else 0,
    )


@app.route("/owners")
@login_required
def owners():
    q            = request.args.get("q", "").strip()
    zip_f        = request.args.get("zip", "").strip()
    muslim_f     = request.args.get("muslim", "").strip()
    status_f     = request.args.get("status", "").strip()
    street_f     = request.args.get("street", "").strip()
    show_ignored = request.args.get("show_ignored", "") == "1"
    page         = max(1, int(request.args.get("page", 1)))
    per_page     = 50

    df    = apply_filters(DF.copy(), q, zip_f, muslim_f, status_f, show_ignored, street_f)
    total = len(df)
    pages = max(1, (total + per_page - 1) // per_page)
    page  = min(page, pages)
    records = enrich(df.iloc[(page-1)*per_page : page*per_page])

    return render_template("owners.html",
        records=records, total=total, page=page, pages=pages, per_page=per_page,
        q=q, zip_f=zip_f, muslim_f=muslim_f, status_f=status_f,
        street_f=street_f, show_ignored=show_ignored,
        zip_options=sorted(DF["zip"].dropna().unique()),
        status_options=STATUS_OPTIONS,
    )


@app.route("/annotate/<rid>", methods=["POST"])
@login_required
def annotate(rid):
    data = request.get_json(force=True)
    for field, value in data.items():
        upsert_annotation(rid, field, value)
    return jsonify(ok=True)


@app.route("/bulk", methods=["POST"])
@login_required
def bulk():
    data  = request.get_json(force=True)
    rids  = data.get("rids", [])
    field = data.get("field", "")
    value = data.get("value", "")
    if rids and field:
        bulk_upsert(rids, field, value)
    return jsonify(ok=True, updated=len(rids))


@app.route("/route")
@login_required
def route():
    """Redirect to Google Maps multi-stop route for a street's Muslim homes."""
    street_f = request.args.get("street", "").strip()
    zip_f    = request.args.get("zip", "").strip()

    df = DF[DF["is_muslim"]].copy()
    if zip_f:
        df = df[df["zip"] == zip_f]
    if street_f:
        df = df[df["street"] == street_f.upper()]

    all_ann = get_all_annotations()
    df = df[df["rid"].apply(lambda r: not all_ann.get(r, {}).get("ignored", 0))]

    addresses = [f"{r['address']}, {r['city_state_zip']}" for r in df.to_dict("records")]
    if not addresses:
        flash("No addresses found for this street.")
        return redirect(url_for("streets"))

    return redirect(make_gmaps_route(addresses))


@app.route("/export")
@login_required
def export():
    q        = request.args.get("q", "").strip()
    zip_f    = request.args.get("zip", "").strip()
    muslim_f = request.args.get("muslim", "").strip()
    status_f = request.args.get("status", "").strip()
    street_f = request.args.get("street", "").strip()
    fmt      = request.args.get("fmt", "csv")

    df  = apply_filters(DF.copy(), q, zip_f, muslim_f, status_f, show_ignored=False, street_f=street_f)
    ann = get_annotations(df["rid"].tolist())

    rows = []
    for rec in df.to_dict("records"):
        a = ann.get(rec["rid"], {})
        rows.append({
            "Last Name":      rec["last_name"],
            "First Name":     rec["first_name"],
            "Address":        rec["address"],
            "Street":         rec["street"],
            "City/State/ZIP": rec["city_state_zip"],
            "Auto-Detected":  "Muslim" if rec["is_muslim"] else "Other",
            "Status":         a.get("status", ""),
            "Last Visited":   a.get("last_visited", ""),
            "Comments":       a.get("comments", ""),
        })

    if fmt == "xlsx":
        return export_xlsx(rows)

    buf = io.StringIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode()),
                     mimetype="text/csv", as_attachment=True,
                     download_name="frisco_export.csv")


def export_xlsx(rows):
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Frisco Export"
    headers = ["Last Name","First Name","Address","Street","City/State/ZIP",
               "Auto-Detected","Status","Last Visited","Comments"]
    widths  = [20, 22, 36, 28, 26, 14, 16, 16, 40]

    ws.append(headers)
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font, c.alignment = HEADER_FILL, HEADER_FONT, Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w

    for rn, r in enumerate(rows, 2):
        ws.append([r[h] for h in headers])
        if rn % 2 == 0:
            for c in ws[rn]:
                c.fill = ALT_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="frisco_export.xlsx")


if __name__ == "__main__":
    print("Starting Frisco Admin at http://127.0.0.1:5000")
    print("Login: admin / admin123")
    app.run(debug=True, port=5000)
