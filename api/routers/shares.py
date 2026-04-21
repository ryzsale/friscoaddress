import io
import json
import secrets
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd

from ..auth import get_current_user
from ..database import get_db
from ..data import get_dataframe
from ..models import Annotation, ShareToken
from ..schemas import ShareCreate, ShareOut
from ..routers.owners import _apply_filters, _enrich
from ..routers.export import _build_rows, HEADERS, WIDTHS

router = APIRouter(prefix="/shares", tags=["shares"])


def _get_valid_token(token: str, db: Session) -> ShareToken:
    t = db.query(ShareToken).filter(ShareToken.token == token).first()
    if not t:
        raise HTTPException(404, "Link not found or expired")
    if datetime.utcnow() > datetime.fromisoformat(t.expires_at):
        raise HTTPException(404, "Link expired")
    return t


@router.post("", response_model=dict)
def create_share(
    body: ShareCreate,
    db: Session = Depends(get_db),
    _: str = Depends(get_current_user),
):
    token = secrets.token_urlsafe(24)
    now   = datetime.utcnow()
    t = ShareToken(
        token      = token,
        filters    = json.dumps(body.filters),
        label      = body.label,
        created_at = now.isoformat(),
        expires_at = (now + timedelta(days=body.days)).isoformat(),
    )
    db.add(t)
    db.commit()
    return {"ok": True, "token": token}


@router.get("", response_model=list[ShareOut])
def list_shares(db: Session = Depends(get_db), _: str = Depends(get_current_user)):
    tokens = db.query(ShareToken).order_by(ShareToken.created_at.desc()).all()
    now    = datetime.utcnow()
    return [ShareOut(
        token=t.token, label=t.label, filters=t.filters,
        created_at=t.created_at, expires_at=t.expires_at,
        expired=now > datetime.fromisoformat(t.expires_at),
    ) for t in tokens]


@router.delete("/{token}")
def delete_share(token: str, db: Session = Depends(get_db), _: str = Depends(get_current_user)):
    db.query(ShareToken).filter(ShareToken.token == token).delete()
    db.commit()
    return {"ok": True}


@router.get("/{token}/data")
def share_data(token: str, db: Session = Depends(get_db)):
    t       = _get_valid_token(token, db)
    filters = json.loads(t.filters)
    df      = _apply_filters(
        get_dataframe().copy(),
        filters.get("q",""), filters.get("zip",""), filters.get("muslim",""),
        filters.get("status",""), False, filters.get("street",""), db,
    )
    return {
        "label":   t.label,
        "total":   len(df),
        "expires": t.expires_at[:10],
        "filters": filters,
        "records": _enrich(df.head(200), db),
    }


@router.get("/{token}/download/csv")
def share_download_csv(token: str, db: Session = Depends(get_db)):
    t       = _get_valid_token(token, db)
    filters = json.loads(t.filters)
    df      = _apply_filters(
        get_dataframe().copy(),
        filters.get("q",""), filters.get("zip",""), filters.get("muslim",""),
        filters.get("status",""), False, filters.get("street",""), db,
    )
    rows = _build_rows(df, db)
    buf  = io.StringIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=frisco_share.csv"})


@router.get("/{token}/download/xlsx")
def share_download_xlsx(token: str, db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    t       = _get_valid_token(token, db)
    filters = json.loads(t.filters)
    df      = _apply_filters(
        get_dataframe().copy(),
        filters.get("q",""), filters.get("zip",""), filters.get("muslim",""),
        filters.get("status",""), False, filters.get("street",""), db,
    )
    rows = _build_rows(df, db)

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")

    wb = Workbook()
    ws = wb.active
    ws.title = "Shared Export"
    ws.append(HEADERS)
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    for rn, r in enumerate(rows, 2):
        ws.append([r[h] for h in HEADERS])
        if rn % 2 == 0:
            for c in ws[rn]:
                c.fill = ALT_FILL
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=frisco_share.xlsx"})
