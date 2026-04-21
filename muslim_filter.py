#!/usr/bin/env python3
"""Filter Frisco owner data for common Muslim names and export to Excel with separate tabs."""

import csv
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Name lists ────────────────────────────────────────────────────────────────

MUSLIM_LAST_NAMES = {
    "khan", "ahmed", "ahmad", "ali", "hassan", "hasan", "hussain", "hussein",
    "husain", "husein", "malik", "sheikh", "shaikh", "shaik", "siddiqui",
    "siddique", "siddiq", "ansari", "qureshi", "quereshi", "chaudhry",
    "chaudhri", "choudhary", "choudhry", "choudry", "rahman", "islam",
    "ullah", "baig", "mirza", "rizvi", "hashmi", "bokhari", "butt",
    "javed", "nawaz", "iqbal", "zafar",
    "raza", "naqvi", "kazmi", "zaidi", "syed", "saeed", "farooq", "usman",
    "awan", "aslam", "bashir", "ghani", "hayat", "jilani", "lodhi", "lodi",
    "memon", "mughal", "niazi", "qazi", "rabbani", "sadiq", "saleem",
    "salim", "suleman", "sultan", "tahir", "wahid", "wasim", "yousuf",
    "yusuf", "zahid", "zaman", "chishti", "dehlvi", "feroze", "ghauri",
    "haider", "hamid", "imran", "ismail", "jabbar", "jamil", "karimi",
    "khalid", "latif", "mahmood", "mahmud", "majid", "maqsood", "masood",
    "mohsin", "mukhtar", "munir", "murad", "nadeem", "naeem", "noman",
    "pervaiz", "pervez", "qadir", "qadri", "rafiq", "rahim", "rashid",
    "rehman", "sajid", "sarwar", "shafiq", "shahid", "shakeel", "sharif",
    "tariq", "waqar", "warsi", "zia", "zubair", "ghazi",
    "mian", "channa", "solangi", "pasha", "bakhsh",
    "al", "el",  # common prefixes (will catch Al-Farsi, El-Amin etc.)
    # Arab last names
    "alfarsi", "alamin", "alansari", "alaziz", "albakri", "alhamad",
    "alhussain", "aljazeera", "alkhalidi", "almansouri", "almasri",
    "almuhairi", "alnajjar", "alomari", "alqahtani", "alrasheedi",
    "alrashidi", "alsaidi", "alshamsi", "alsuwaidi", "alzahrani",
    "bin", "bint",
    # Turkish / Central Asian
    "yilmaz", "kaya", "demir", "celik", "sahin", "yildiz", "ozturk",
    "arslan", "ozdemir", "kaplan",
    # Iranian / Persian
    "hosseini", "hosseinzadeh", "mohammadi", "ahmadi", "rezaei", "mousavi",
    "sadeghi", "karimi", "tehrani", "shirazi", "isfahani", "ghafari",
    "sadri", "ghiassi", "jafari", "nazari", "ebrahimi",
    "rashidian", "ardestani", "kashani",
    # Indonesian / Malay
    "hidayat",
}

MUSLIM_FIRST_NAMES = {
    # Male
    "muhammad", "mohammed", "mohammad", "mehmet", "mohamad", "mohamed",
    "ahmed", "ahmad", "ali", "omar", "umar", "ibrahim", "yusuf", "yousef",
    "younis", "yunus", "khalid", "bilal", "hamza", "hamzah", "karim",
    "kareem", "tariq", "tarik", "faisal", "zaid", "zayd", "imran",
    "amir", "ameer", "mustafa", "moustafa", "abdullah", "abdulla",
    "abdulrahman", "abdul", "syed", "sayyid", "hasan", "hussain",
    "hussein", "raza", "rashid", "raashid", "farrukh", "farhan",
    "farid", "faris", "danish", "daniyar", "dawood", "dawud",
    "idris", "isa", "issa", "jabir", "jafar", "jalal", "jamil",
    "javid", "jibril", "kamran", "kashif", "khurram", "luqman",
    "mahir", "mahmood", "mahmud", "majid", "mansoor", "mansur",
    "marwan", "masood", "masud", "mubarak", "mubin", "mukhtar",
    "munir", "murad", "musa", "musab", "nabeel", "nabil", "nadeem",
    "naeem", "nasr", "nasser", "nasir", "nawaz", "noman", "nouman",
    "obaid", "osama", "osman", "qasim", "raees", "rafiq", "rahim",
    "ramadan", "rizwan", "sadiq", "saeed", "said", "sajid", "saleh",
    "salim", "salman", "sarwar", "shahid", "shahrukh", "shakeel",
    "shaukat", "sulaiman", "suleiman", "tabish", "taha", "tahir",
    "talha", "talhah", "usman", "waheed", "wahid", "waleed", "walid",
    "waseem", "wasim", "waqar", "yasin", "yaseen", "yasir", "zahid",
    "zahir", "zaid", "zaki", "zeeshan", "zubair", "aamir", "adeel",
    "adil", "adnan", "ahsan", "aijaz", "akbar", "akram", "alam",
    "aleem", "anwar", "aqeel", "arif", "arsalan", "arshad", "asad",
    "asif", "asim", "atif", "awais", "ayaz", "azhar", "aziz",
    "babar", "basit", "bilal", "burhan", "daniyal", "ghulam",
    "habib", "hafiz", "hamid", "haroon", "haroun", "harun",
    # Female
    "fatima", "fatimah", "aisha", "ayesha", "aysha", "maryam",
    "mariam", "zainab", "zaynab", "zahra", "nadia",
    "nadya", "layla", "amira", "ameera", "noor", "nur",
    "sana", "yasmin", "yasmine", "samira", "khadijah",
    "khadija", "hafsa", "hafiza", "ruqayyah", "sumayyah",
    "asma", "asmaa", "rania", "rana",
    "huda", "iman", "lubna", "maha", "manal", "marwa",
    "najwa", "nawal", "nour", "reem", "rima", "shaima",
    "shaheen", "shahina", "shahida", "shamsa", "shirin", "suraya",
    "tabassum", "tahira", "zara", "zuhra", "zulaikha",
    "bushra", "farida", "fauzia", "fiza", "fozia", "ghazala",
    "gulnaz", "haleema", "halima", "humera", "isra", "khola",
    "madiha", "maheen", "mahira", "maliha", "munira",
    "nafeesa", "nargis", "nasreen", "nazia", "nilofer", "parveen",
    "rabia", "rabiya", "rahila", "razia", "rukhsana", "saima",
    "sajida", "sakeena", "sakina", "sanam", "sehrish", "shagufta",
    "shaista", "shamim", "sidra", "soha", "suha", "sundus",
    "umm", "uzma", "zobia", "zohra",
}


def is_muslim_name(last: str, first: str) -> bool:
    l = last.lower().strip()
    f = first.lower().strip().split()[0] if first.strip() else ""

    if l in MUSLIM_LAST_NAMES:
        return True
    # catch hyphenated or prefixed Arab names: Al-Rashidi, El-Amin
    if l.startswith("al-") or l.startswith("el-"):
        return True
    if f in MUSLIM_FIRST_NAMES:
        return True
    return False


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
COLS        = ["last_name", "first_name", "address", "city_state_zip"]
COL_WIDTHS  = [20, 22, 35, 25]


def write_sheet(ws, rows: list[dict]) -> None:
    ws.append(["Last Name", "First Name", "Address", "City / State / ZIP"])
    for col_idx, (width, cell) in enumerate(
        zip(COL_WIDTHS, ws[1]), start=1
    ):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_num, rec in enumerate(rows, start=2):
        ws.append([rec["last_name"], rec["first_name"], rec["address"], rec["city_state_zip"]])
        if row_num % 2 == 0:
            for cell in ws[row_num]:
                cell.fill = ALT_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Filter Muslim names from Frisco CSV and export to Excel."
    )
    parser.add_argument("input_csv", help="CSV file produced by extract.py")
    parser.add_argument("-o", "--output", help="Output Excel file (default: muslim_names.xlsx)")
    parser.add_argument(
        "--min-tab", type=int, default=3,
        help="Minimum records for a last name to get its own tab (default: 3)"
    )
    args = parser.parse_args()
    output = args.output or "muslim_names.xlsx"

    # ── Read & filter ────────────────────────────────────────────────────────
    all_matches: list[dict] = []
    by_last: dict[str, list[dict]] = defaultdict(list)

    with open(args.input_csv, newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            if is_muslim_name(rec["last_name"], rec["first_name"]):
                all_matches.append(rec)
                by_last[rec["last_name"].title()].append(rec)

    print(f"Found {len(all_matches)} records with Muslim names "
          f"across {len(by_last)} distinct last names.")

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = Workbook()

    # Summary tab
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Last Name", "Count"])
    for cell in ws_sum[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 10

    sorted_names = sorted(by_last.items(), key=lambda x: -len(x[1]))
    for name, recs in sorted_names:
        ws_sum.append([name, len(recs)])
    ws_sum.auto_filter.ref = "A1:B1"
    ws_sum.freeze_panes = "A2"

    # "All" tab
    ws_all = wb.create_sheet("All")
    write_sheet(ws_all, all_matches)

    # One tab per last name with enough records
    for name, recs in sorted_names:
        if len(recs) < args.min_tab:
            continue
        safe_title = name[:31]  # Excel sheet name limit
        ws = wb.create_sheet(safe_title)
        write_sheet(ws, recs)

    wb.save(output)
    print(f"Saved → {output}")
    print(f"  • Summary tab: {len(by_last)} last names")
    print(f"  • All tab: {len(all_matches)} records")
    tabs = sum(1 for _, r in sorted_names if len(r) >= args.min_tab)
    print(f"  • Individual tabs: {tabs} last names with ≥{args.min_tab} records")


if __name__ == "__main__":
    main()
